import json
import requests
import qrcode
import base64

from io import BytesIO
from base64 import b64encode
from datetime import date, time, timedelta, datetime
import os
import mimetypes
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
from django.urls import reverse
from django.core.paginator import Paginator
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required, user_passes_test
from django.db import transaction
from django.db.models import Q, Exists, OuterRef, Case, When, IntegerField
from django.contrib.auth.models import Group

from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.enums import TA_CENTER
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from .forms import EmpleadoForm
from .models import User_Empleados
from Core_session.models import HuellaCaptura
from Asistencia.models import Asistencia
from Membresias.models import Membresia
from Rutinas.models import Rutina
from Historial.utils import registrar_movimiento


@permission_required("Empleados.view_user_empleados", raise_exception=True)
@transaction.atomic
def ListarEmpleados(request):
    
    # Verificar si hay mensaje pendiente de la creación
    mensaje_exito = request.session.pop('mensaje_exito', None)
    if mensaje_exito:
        messages.success(request, mensaje_exito)
        
    # Obtener parámetros de búsqueda, filtro y paginación
    search_query = request.GET.get("search", "").strip()
    filter_state = request.GET.get("estado", "").strip()
    filter_group = request.GET.get("rol", "").strip()
    page_number = request.GET.get("page", 1)
    items_per_page = int(request.GET.get("items_per_page", 10))

    empleados = (
        User_Empleados.objects
        .annotate(tiene_huella=Exists(
            HuellaCaptura.objects.filter(id_usuario=OuterRef('pk'))
        ))
        .order_by('id')
    )

    if request.user.groups.filter(name="Empleados").exists():
        empleados = empleados.filter(groups__name="Usuarios")

    # Búsqueda por nombre, apellido, correo o cédula
    if search_query:
        search_filters = (
            Q(first_name__icontains=search_query)
            | Q(last_name__icontains=search_query)
            | Q(email__icontains=search_query)
        )
        if search_query.isdigit():
            search_filters |= Q(Cedula=int(search_query))
        empleados = empleados.filter(search_filters)

    if filter_state == "active":
        empleados = empleados.filter(is_active=True)
    elif filter_state == "inactive":
        empleados = empleados.filter(is_active=False)

    if filter_group.startswith("group_"):
        group_id = filter_group.replace("group_", "")
        try:
            empleados = empleados.filter(groups__id=group_id).distinct()
        except ValueError:
            pass

    paginator = Paginator(empleados, items_per_page)
    page_obj = paginator.get_page(page_number)
    pagina_ids = ",".join(str(emp.id) for emp in page_obj.object_list)
    pagination_params = request.GET.copy()
    pagination_params.pop("page", None)
    pagination_suffix = pagination_params.urlencode()
    if pagination_suffix:
        pagination_suffix = "&" + pagination_suffix

    context = {
        "Empleados": page_obj,
        "total_items": paginator.count,
        "search_query": search_query,
        "filter_state": filter_state,
        "filter_group": filter_group,
        "items_per_page": items_per_page,
        "grupos_disponibles": Group.objects.all().order_by("name"),
        "puede_ver_grupos": request.user.is_superuser or request.user.groups.filter(name="Admin").exists(),
        "pagina_ids": pagina_ids,
        "pagination_suffix": pagination_suffix,
    }
    return render(request, "templates_usuarios/usuarios.html", context)


def _obtener_empleados_para_export(ids):
    ordering = Case(
        *[When(id=pk, then=pos) for pos, pk in enumerate(ids)],
        default=len(ids),
        output_field=IntegerField(),
    )

    return list(
        User_Empleados.objects.filter(id__in=ids)
        .prefetch_related("groups")
        .annotate(_order=ordering)
        .order_by("_order")
    )


@login_required(login_url='login')
@permission_required("Empleados.view_user_empleados", raise_exception=True)
def ExportarUsuariosPDF(request):
    """Genera un PDF con los usuarios visibles actualmente en la tabla."""
    if request.method != "POST":
        messages.error(request, "Selecciona los usuarios que deseas exportar.")
        return redirect("Empleados")

    ids_raw = request.POST.get("visible_ids", "").strip()
    if not ids_raw:
        messages.error(request, "No hay usuarios para exportar.")
        return redirect("Empleados")

    try:
        ids = [int(value) for value in ids_raw.split(",") if value.strip().isdigit()]
    except ValueError:
        ids = []

    if not ids:
        messages.error(request, "No se recibieron usuarios válidos para el reporte.")
        return redirect("Empleados")

    empleados = _obtener_empleados_para_export(ids)

    if not empleados:
        messages.error(request, "No se encontraron usuarios para exportar.")
        return redirect("Empleados")

    activos = sum(1 for e in empleados if e.is_active)
    inactivos = len(empleados) - activos

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=45,
        leftMargin=45,
        topMargin=50,
        bottomMargin=50,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "UsuariosTitle",
        parent=styles["Heading1"],
        fontSize=20,
        textColor=colors.HexColor("#8B0000"),
        alignment=TA_CENTER,
        spaceAfter=8,
    )
    subtitle_style = ParagraphStyle(
        "UsuariosSubtitle",
        parent=styles["Heading2"],
        fontSize=13,
        alignment=TA_CENTER,
        spaceAfter=20,
    )
    normal_style = ParagraphStyle(
        "UsuariosNormal",
        parent=styles["Normal"],
        fontSize=10,
        spaceAfter=4,
    )

    elements = [
        Paragraph("ATHLETIC-Q GIMNASIO", title_style),
        Paragraph("Reporte de Usuarios", subtitle_style),
        Paragraph(
            f"<b>Generado por:</b> {request.user.get_full_name() or request.user.username}",
            normal_style,
        ),
        Paragraph(
            f"<b>Fecha:</b> {timezone.now().strftime('%d/%m/%Y %H:%M')}",
            normal_style,
        ),
        Spacer(1, 12),
    ]

    def _build_photo_cell(usuario):
        if usuario.empleados_img and getattr(usuario.empleados_img, "path", None):
            try:
                return Image(usuario.empleados_img.path, width=28, height=28)
            except Exception:
                pass
        return Paragraph("—", normal_style)

    table_data = [["#", "Foto", "Cédula", "Nombre", "Correo", "Rol", "Estado"]]
    for idx, empleado in enumerate(empleados, 1):
        roles = ", ".join(empleado.groups.values_list("name", flat=True)) or "Sin rol"
        estado = "Activo" if empleado.is_active else "Inactivo"
        table_data.append(
            [
                str(idx),
                _build_photo_cell(empleado),
                empleado.Cedula or "—",
                f"{empleado.first_name} {empleado.last_name}".strip() or empleado.username,
                empleado.email or "—",
                roles,
                estado,
            ]
        )

    table = Table(table_data, colWidths=[25, 45, 70, 120, 130, 110, 60])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#8B0000")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 10),
                ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 1), (-1, -1), 9),
                ("ALIGN", (0, 1), (0, -1), "CENTER"),
                ("ALIGN", (1, 1), (1, -1), "CENTER"),
                ("ALIGN", (5, 1), (5, -1), "CENTER"),
                ("ALIGN", (6, 1), (6, -1), "CENTER"),
                ("LINEBELOW", (0, 0), (-1, 0), 1, colors.grey),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#fafafa")]),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )
    elements.append(table)
    elements.append(Spacer(1, 18))

    resumen = [
        ["Usuarios incluidos:", str(len(empleados))],
        ["Activos:", str(activos)],
        ["Inactivos:", str(inactivos)],
    ]
    resumen_table = Table(resumen, colWidths=[140, 70])
    resumen_table.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#444444")),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    elements.append(resumen_table)
    elements.append(Spacer(1, 20))

    footer_style = ParagraphStyle(
        "UsuariosFooter",
        parent=styles["Normal"],
        fontSize=8,
        textColor=colors.grey,
        alignment=TA_CENTER,
    )
    elements.append(Paragraph("──────────────────────────────────────────────", footer_style))
    elements.append(Spacer(1, 6))
    elements.append(
        Paragraph("Reporte generado por Athletic-Q | Gestión de Usuarios", footer_style)
    )

    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = "attachment; filename=usuarios_filtrados.pdf"
    response.write(pdf)
    return response


@login_required(login_url='login')
@permission_required("Empleados.view_user_empleados", raise_exception=True)
def ExportarUsuariosExcel(request):
    if request.method != "POST":
        messages.error(request, "Selecciona los usuarios que deseas exportar.")
        return redirect("Empleados")

    ids_raw = request.POST.get("visible_ids", "").strip()
    if not ids_raw:
        messages.error(request, "No hay usuarios para exportar.")
        return redirect("Empleados")

    ids = [int(value) for value in ids_raw.split(",") if value.strip().isdigit()]
    if not ids:
        messages.error(request, "No se recibieron usuarios válidos para el reporte.")
        return redirect("Empleados")

    empleados = _obtener_empleados_para_export(ids)
    if not empleados:
        messages.error(request, "No se encontraron usuarios para exportar.")
        return redirect("Empleados")

    activos = sum(1 for e in empleados if e.is_active)
    inactivos = len(empleados) - activos

    wb = Workbook()
    ws = wb.active
    ws.title = "Usuarios"

    # Ajustar anchos de columna
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 9
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 12

    # Encabezado principal
    ws.merge_cells("A1:G1")
    ws["A1"] = "ATHLETIC-Q GIMNASIO"
    ws["A1"].font = Font(size=18, bold=True, color="8B0000")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A2:G2")
    ws["A2"] = "Reporte de Usuarios"
    ws["A2"].font = Font(size=13, bold=True)
    ws["A2"].alignment = Alignment(horizontal="center")

    # Información general
    info_rows = [
        ("Generado por:", request.user.get_full_name() or request.user.username, "Fecha:", timezone.now().strftime('%d/%m/%Y %H:%M')),
        ("Usuarios incluidos:", len(empleados), "Activos:", activos),
        ("Inactivos:", inactivos, "", ""),
    ]

    start_row = 4
    for left_label, left_value, right_label, right_value in info_rows:
        ws[f"A{start_row}"] = left_label
        ws[f"A{start_row}"].font = Font(bold=True)
        ws[f"B{start_row}"] = left_value

        if right_label:
            ws[f"D{start_row}"] = right_label
            ws[f"D{start_row}"].font = Font(bold=True)
            ws[f"E{start_row}"] = right_value
        start_row += 1

    header_row = start_row + 1
    headers = ["#", "Foto", "Cédula", "Nombre", "Correo", "Rol", "Estado"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="8B0000")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    thin_border = Border(
        left=Side(style="thin", color="4B4B4B"),
        right=Side(style="thin", color="4B4B4B"),
        top=Side(style="thin", color="4B4B4B"),
        bottom=Side(style="thin", color="4B4B4B"),
    )

    current_row = header_row + 1
    for idx, empleado in enumerate(empleados, 1):
        roles = ", ".join(empleado.groups.values_list("name", flat=True)) or "Sin rol"
        estado = "Activo" if empleado.is_active else "Inactivo"

        data = [
            idx,
            "",
            empleado.Cedula or "—",
            f"{empleado.first_name} {empleado.last_name}".strip() or empleado.username,
            empleado.email or "—",
            roles,
            estado,
        ]

        for col, value in enumerate(data, 1):
            cell = ws.cell(row=current_row, column=col, value=value)
            cell.border = thin_border
            if col in (1, 7):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(vertical="center")

        ws.row_dimensions[current_row].height = 34

        foto_path = (
            empleado.empleados_img.path
            if empleado.empleados_img and getattr(empleado.empleados_img, "path", None)
            else None
        )
        if foto_path:
            ext = os.path.splitext(foto_path)[1].lower()
            if ext not in {".png", ".jpg", ".jpeg"}:
                foto_path = None

        if foto_path:
            try:
                xl_image = XLImage(foto_path)
                xl_image.width = 32
                xl_image.height = 32
                xl_image.anchor = f"B{current_row}"
                ws.add_image(xl_image)
            except Exception:
                pass

        current_row += 1

    summary_start = current_row + 1
    ws.cell(row=summary_start, column=5, value="TOTAL USUARIOS:").font = Font(bold=True)
    ws.cell(row=summary_start, column=6, value=len(empleados))
    ws.cell(row=summary_start + 1, column=5, value="ACTIVOS:").font = Font(bold=True)
    ws.cell(row=summary_start + 1, column=6, value=activos)
    ws.cell(row=summary_start + 2, column=5, value="INACTIVOS:").font = Font(bold=True)
    ws.cell(row=summary_start + 2, column=6, value=inactivos)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=usuarios_filtrados.xlsx"
    return response


@transaction.atomic
def CrearEmpleado(request):
    if request.method == "POST":
        form = EmpleadoForm(request.POST, request.FILES, usuario_actual=request.user)

        if form.is_valid():
            try:
                empleado = form.save(commit=False)

                # Flags requeridos por el modelo/auth
                empleado.is_superuser = False
                empleado.is_staff = True
                is_active_raw = form.cleaned_data.get("is_active")
                empleado.is_active = str(is_active_raw).lower() in ("true", "1", "yes", "on", "activo")

                # Asegurar fecha de creación para no enviar NULL
                if not empleado.date_joined:
                    empleado.date_joined = timezone.now()

                # Guardar contraseña correctamente
                empleado.set_password(form.cleaned_data["password"])
                empleado.save()

                # Guardar grupo/rol
                if form.cleaned_data.get("groups"):
                    empleado.groups.add(form.cleaned_data["groups"])

                nombre_objeto = f"{empleado.first_name} {empleado.last_name}".strip() or empleado.username
                registrar_movimiento(
                    user=request.user,
                    tipo="ingresar",
                    modulo="usuarios",
                    nombre_objeto=nombre_objeto,
                    id_objeto=empleado.id,
                )

                # Guardar mensaje en sesión para mostrar DESPUÉS del modal
                request.session['mensaje_exito'] = f'Usuario "{empleado.first_name} {empleado.last_name}" creado exitosamente.'

                url = f"{reverse('empleados_create')}?show_modal=1&emp_id={empleado.id}"
                return redirect(url)

            except Exception as e:
                transaction.set_rollback(True)
                messages.error(request, f"Error al crear empleado: {e}")
                return render(request, "templates_usuarios/crear_usuarios.html", {"form": form})

        messages.error(request, 'Error en el formulario. Revisa los campos marcados.')
        return render(request, "templates_usuarios/crear_usuarios.html", {"form": form})

    # GET: mostrar el formulario
    form = EmpleadoForm(usuario_actual=request.user)
    return render(request, "templates_usuarios/crear_usuarios.html", {"form": form})

@transaction.atomic
def EditarEmpleado(request, id):
    empleado = get_object_or_404(User_Empleados, id=id)
    old_password = empleado.password  # guardamos la contraseña actual (hash)
    
    if request.method == "POST":
        form = EmpleadoForm(request.POST, request.FILES, instance=empleado, usuario_actual=request.user)
        
        if form.is_valid():
            empleado = form.save(commit=False)

            # ⚠️ Esto viene de clean_password (lo ajustamos abajo)
            password = form.cleaned_data.get('password', '').strip()

            # Si el usuario escribió una nueva contraseña (distinta del hash anterior),
            # la seteamos. Si viene igual que old_password, asumimos "no cambio".
            if password and password != old_password:
                empleado.set_password(password)

            empleado.save()
            
            # Actualizar grupos
            if form.cleaned_data.get('groups'):
                empleado.groups.set([form.cleaned_data.get('groups')])

            nombre_objeto = f"{empleado.first_name} {empleado.last_name}".strip() or empleado.username
            registrar_movimiento(
                user=request.user,
                tipo="editar",
                modulo="usuarios",
                nombre_objeto=nombre_objeto,
                id_objeto=empleado.id,
            )
            
            messages.success(request, f"Empleado '{empleado.username}' actualizado exitosamente.")
            return redirect('empleados_detail', id=empleado.id)
        else:
            messages.error(request, "Hay errores en el formulario. Por favor, revísalo.")
    else:
        form = EmpleadoForm(instance=empleado, usuario_actual=request.user)
    
    context = {
        'form': form,
        'empleado': empleado
    }
    
    return render(request, "templates_usuarios/editar_usuarios.html", context)

@login_required(login_url='login')
@permission_required("Empleados.view_user_empleados", raise_exception=True)
def DetalleEmpleado(request, id):
    """Vista para ver el detalle completo de un usuario"""
    empleado = get_object_or_404(
        User_Empleados.objects.prefetch_related('groups'),
        id=id
    )
    
    # Verificar si tiene huella registrada
    tiene_huella = HuellaCaptura.objects.filter(id_usuario=empleado).exists()
    
    # Obtener membresía activa (si existe)
    membresia_activa = Membresia.objects.filter(
        id_usuario=empleado,
        Estado='Activo'
    ).select_related('For_Id_tipo_membresia').first()
    
    # Contar asistencias del mes actual
    from datetime import date
    hoy = date.today()
    asistencias_mes = Asistencia.objects.filter(
        id_usuario=empleado,
        fecha_entrada__year=hoy.year,
        fecha_entrada__month=hoy.month
    ).count()
    
    context = {
        'empleado': empleado,
        'tiene_huella': tiene_huella,
        'membresia_activa': membresia_activa,
        'asistencias_mes': asistencias_mes,
    }
    
    return render(request, 'templates_usuarios/detalle_usuarios.html', context)

def capturar_huella(request, empleado_id):
    empleado = get_object_or_404(User_Empleados, id=empleado_id)
    return render(request, "templates_huella/huella.html", {"empleado": empleado})

@csrf_exempt
def guardar_huellas(request):
    """Guardar huella capturada en la base de datos"""
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'Método no permitido'}, status=405)
    
    try:
        import json
        
        data = json.loads(request.body)
        usuario_id = data.get('usuario_id')
        templates = data.get('templates', [])
        
        if not usuario_id or not templates:
            return JsonResponse({'ok': False, 'error': 'Datos incompletos'}, status=400)
        
        template = templates[0] if templates else None
        
        if not template:
            return JsonResponse({'ok': False, 'error': 'Template vacío'}, status=400)
        
        # =============================================
        # VERIFICAR DUPLICADOS DIRECTAMENTE EN LA BD
        # (Sin conectarse al servidor de huellas)
        # =============================================
        huella_duplicada = HuellaCaptura.objects.filter(
            template=template
        ).exclude(
            id_usuario_id=usuario_id
        ).first()
        
        if huella_duplicada:
            return JsonResponse({
                'ok': False,
                'error': f"Esta huella ya está registrada para otro usuario (ID: {huella_duplicada.id_usuario_id})"
            }, status=409)
        
        # Obtener usuario
        usuario = get_object_or_404(User_Empleados, id=usuario_id)
        
        # Verificar si ya tiene huella y actualizarla o crear nueva
        huella_existente = HuellaCaptura.objects.filter(id_usuario=usuario).first()
        
        if huella_existente:
            huella_existente.template = template
            huella_existente.save()
            mensaje = 'Huella actualizada correctamente'
        else:
            HuellaCaptura.objects.create(
                id_usuario=usuario,
                template=template
            )
            mensaje = 'Huella guardada correctamente'
        
        return JsonResponse({'ok': True, 'message': mensaje})
        
    except json.JSONDecodeError:
        return JsonResponse({'ok': False, 'error': 'JSON inválido'}, status=400)
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=500)

@csrf_exempt
def obtener_huellas(request):
    """Retorna todas las huellas registradas para comparación local."""
    if request.method != 'GET':
        return JsonResponse({'ok': False, 'error': 'Método no permitido'}, status=405)
    
    try:
        huellas = HuellaCaptura.objects.all().select_related('id_usuario')
        
        huellas_list = []
        for huella in huellas:
            if huella.template and huella.id_usuario:
                huellas_list.append({
                    'usuario_id': huella.id_usuario.id,
                    'nombre': f"{huella.id_usuario.first_name} {huella.id_usuario.last_name}",
                    'template': huella.template
                })
        
        return JsonResponse({
            'ok': True,
            'huellas': huellas_list,
            'total': len(huellas_list)
        })
        
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=500)


@csrf_exempt
def validar_huella(request):
    """
    Recibe el usuario_id del match hecho en el frontend.
    NO se conecta a localhost:5000.
    """
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'Método no permitido'}, status=405)
    
    try:
        import json
        
        data = json.loads(request.body)
        usuario_id = data.get('usuario_id')
        
        if not usuario_id:
            return JsonResponse({
                'ok': False,
                'error': 'No se recibió usuario_id'
            }, status=400)
        
        # Buscar usuario
        usuario = get_object_or_404(User_Empleados, id=usuario_id)
        
        # =============================================
        # REGISTRAR ASISTENCIA
        # =============================================
        hoy = timezone.now()
        inicio_dia = hoy.replace(hour=0, minute=0, second=0, microsecond=0)
        fin_dia = hoy.replace(hour=23, minute=59, second=59, microsecond=999999)
        
        asistencia_abierta = Asistencia.objects.filter(
            id_usuario=usuario,
            fecha_entrada__range=(inicio_dia, fin_dia),
            fecha_salida__isnull=True
        ).first()
        
        if asistencia_abierta:
            # Registrar SALIDA
            asistencia_abierta.fecha_salida = hoy
            asistencia_abierta.estado = 'Completado'
            asistencia_abierta.save()
            
            return JsonResponse({
                'ok': True,
                'action': 'salida',
                'message': 'Salida registrada correctamente',
                'usuario': {
                    'id': usuario.id,
                    'nombre': f"{usuario.first_name} {usuario.last_name}",
                    'username': usuario.username
                },
                'asistencia': {
                    'id': asistencia_abierta.id,
                    'entrada': asistencia_abierta.fecha_entrada.strftime('%Y-%m-%d %H:%M:%S'),
                    'salida': hoy.strftime('%Y-%m-%d %H:%M:%S'),
                    'estado': 'Completado'
                }
            })
        else:
            # Registrar ENTRADA
            grupo = usuario.groups.first()
            rol = grupo.name if grupo else 'Sin rol'
            
            nueva_asistencia = Asistencia.objects.create(
                id_usuario=usuario,
                rol=rol,
                fecha_entrada=hoy,
                estado='Pendiente'
            )
            
            return JsonResponse({
                'ok': True,
                'action': 'entrada',
                'message': 'Entrada registrada correctamente',
                'usuario': {
                    'id': usuario.id,
                    'nombre': f"{usuario.first_name} {usuario.last_name}",
                    'username': usuario.username
                },
                'asistencia': {
                    'id': nueva_asistencia.id,
                    'entrada': hoy.strftime('%Y-%m-%d %H:%M:%S'),
                    'salida': None,
                    'estado': 'Pendiente'
                }
            })
            
    except json.JSONDecodeError:
        return JsonResponse({'ok': False, 'error': 'JSON inválido'}, status=400)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'ok': False, 'error': str(e)}, status=500)

@login_required(login_url="login")
def MiPerfil(request):
    """Vista de perfil para administradores y empleados"""
    usuario = request.user

    qr_image = generar_qr_asistencia(
        request,
        usuario_id=usuario.id,
        usuario_nombre=usuario.first_name or usuario.username
    )

    membresias = Membresia.objects.filter(id_usuario=usuario).order_by("-Fecha_inicio")
    grupos = usuario.groups.all()
    tiene_huella = HuellaCaptura.objects.filter(id_usuario=usuario).exists()

    context = {
        "usuario": usuario,
        "membresias": membresias,
        'qr_image': qr_image,
        "grupos": grupos,
        "tiene_huella": tiene_huella,
    }
    return render(request, "templates_perfil/mi_perfil.html",context)

# Función helper para verificar si el usuario pertenece al grupo "Usuarios"
def es_usuario_gym(user):
    return user.groups.filter(name='Usuarios').exists()

# @permission_required('Empleados.view_suariogym', login_url='home') este permiso puede hacer todo el view si ese usuario tiene ese permiso
@user_passes_test(es_usuario_gym, login_url='home')
def UsersGym(request): 
    """Vista principal del cliente/usuario del gimnasio"""
    usuario = request.user

    # --------- LÓGICA DE QR (versión 1) ----------
    user = usuario
    membresia = Membresia.objects.filter(id_usuario=user).first()

    qr_image = generar_qr_asistencia(
        request,
        usuario_id=user.id,
        usuario_nombre=user.first_name or user.username
    )
    # --------------------------------------------

    # Obtener membresía activa del usuario (versión 2)
    membresia_activa = Membresia.objects.filter(
        id_usuario=usuario,
        Estado='Activo'
    ).select_related('For_Id_tipo_membresia').first()
    
    # Calcular días restantes si hay membresía
    dias_restantes = 0
    progreso = 0
    if membresia_activa:
        from datetime import date
        hoy = date.today()
        if membresia_activa.Fecha_fin >= hoy:
            dias_restantes = (membresia_activa.Fecha_fin - hoy).days
            # Calcular progreso
            fecha_inicio = membresia_activa.Fecha_inicio.date() if hasattr(membresia_activa.Fecha_inicio, 'date') else membresia_activa.Fecha_inicio
            total_dias = (membresia_activa.Fecha_fin - fecha_inicio).days
            dias_transcurridos = (hoy - fecha_inicio).days
            progreso = min(100, max(0, (dias_transcurridos / total_dias) * 100)) if total_dias > 0 else 0
    
    # Obtener rutinas activas
    rutinas = Rutina.objects.filter(Estado='Activo').order_by('Categoria', 'Nivel')
    
    # Agrupar rutinas por categoría
    rutinas_por_categoria = {}
    for rutina in rutinas:
        if rutina.Categoria not in rutinas_por_categoria:
            rutinas_por_categoria[rutina.Categoria] = []
        rutinas_por_categoria[rutina.Categoria].append(rutina)
    
    context = {
        'usuario': usuario,
        'membresia': membresia_activa,          # membresía activa (versión 2)
        'dias_restantes': dias_restantes,
        'progreso': round(progreso, 1),
        'rutinas': rutinas,
        'rutinas_por_categoria': rutinas_por_categoria,

        # claves de la versión 1 para no romper nada:
        'membresias': membresia,                # la primera que encontraba, sin filtrar por Estado
        'qr_image': qr_image,
    }
    
    return render(request, 'templates_perfil/datos.html', context)

def generar_qr_asistencia(request, usuario_id, usuario_nombre):
    # URL ABSOLUTA 
    url_qr = request.build_absolute_uri(
        reverse('validar_qr', args=[usuario_id])  # sin namespace si no lo usas
    )

    qr = qrcode.QRCode(version=1, box_size=10, border=4)
    qr.add_data(url_qr)
    qr.make(fit=True)
    img = qr.make_image(fill='black', back_color='white')

    buffer = BytesIO()
    img.save(buffer, format='PNG')
    buffer.seek(0)

    qr_image_base64 = base64.b64encode(buffer.getvalue()).decode('utf-8')
    return qr_image_base64

def validar_qr(request, usuario_id):
    usuario = get_object_or_404(User_Empleados, id=usuario_id)

    return render(
        request,
        "templates_perfil/validar_qr.html",
        {"usuario":usuario}
    )

@csrf_exempt
def registrar_asistencia_ajax(request):
    if request.method != "POST":
        return JsonResponse({"ok": False, "mensaje": "Solo POST"}, status=405)

    WAIT_SECONDS = 20

    try:
        data = json.loads(request.body or "{}")
    except json.JSONDecodeError:
        return JsonResponse({"ok": False, "mensaje": "JSON inválido"}, status=400)

    raw = data.get("dato_qr")
    if not raw:
        return JsonResponse({"ok": False, "mensaje": "ID no recibido"}, status=400)

    # Si viene una URL, tomar el último segmento
    if isinstance(raw, str) and "/" in raw:
        raw = raw.strip("/").split("/")[-1]

    try:
        usuario_id = int(raw)
    except (TypeError, ValueError):
        return JsonResponse({"ok": False, "mensaje": "QR inválido"}, status=400)

    try:
        usuario = User_Empleados.objects.get(id=usuario_id)
    except User_Empleados.DoesNotExist:
        return JsonResponse({"ok": False, "mensaje": "QR inválido"}, status=404)

    if not usuario.is_active:
        return JsonResponse({"ok": False, "mensaje": "QR inválido"}, status=400)

    def format_dt(valor):
        if not valor:
            return None
        return timezone.localtime(valor).strftime("%Y-%m-%d %H:%M:%S")

    hoy = timezone.localdate()
    ahora = timezone.now()
    hora_local = timezone.localtime(ahora).time()

    # Ventana horaria permitida: 5:00 - 22:00 (10 p.m.)
    if not (time(1, 0) <= hora_local < time(22, 0)):
        return JsonResponse({"ok": False, "mensaje": "Registro fuera de horario (5:00 a 22:00)."})

    # Último registro abierto del usuario
    reg = Asistencia.objects.filter(
        id_usuario=usuario,
        fecha_salida__isnull=True
    ).order_by('-fecha_entrada').first()

    # Si tiene un registro sin salida de días anteriores, cerrarlo como pendiente
    if reg and reg.fecha_entrada.date() < hoy:
        cierre_prev = datetime.combine(reg.fecha_entrada.date(), time(22, 0))
        if timezone.is_naive(cierre_prev):
            cierre_prev = timezone.make_aware(cierre_prev, timezone.get_current_timezone())
        reg.fecha_salida = cierre_prev
        reg.estado = "Pendiente"
        reg.save(update_fields=["fecha_salida", "estado"])
        reg = None

    # Evita duplicar si acaba de marcar hace pocos segundos
    if reg and (ahora - reg.fecha_entrada) < timedelta(seconds=WAIT_SECONDS):
        mensaje_asistencia = "Debe esperar 20 segundos para registrar la salida."
        accion = "espera"
    else:
        if reg and reg.fecha_entrada.date() == hoy:
            # cerrar salida
            reg.fecha_salida = ahora
            reg.estado = "Salida"
            reg.save(update_fields=["fecha_salida", "estado"])
            mensaje_asistencia = "Salida registrada"
            accion = "salida"
        else:
            # crear nueva entrada
            reg = Asistencia.objects.create(
                id_usuario=usuario,
                rol=(usuario.groups.first().name if usuario.groups.exists() else ""),
                fecha_entrada=ahora,
                fecha_salida=None,
                estado="Entrada"
            )
            mensaje_asistencia = "Entrada registrada"
            accion = "entrada"

    nombre = f"{usuario.first_name} {usuario.last_name}".strip() or usuario.username or f"ID {usuario.id}"
    mensaje_full = (
        mensaje_asistencia
        if accion == "espera"
        else f"Asistencia validada. La asistencia del usuario {nombre} ha sido registrada correctamente."
    )

    return JsonResponse({
        "ok": True,
        "action": accion,
        "usuario": {
            "id": usuario.id,
            "nombre": nombre,
            "username": usuario.username
        },
        "asistencia": {
            "estado": reg.estado,
            "entrada": format_dt(reg.fecha_entrada),
            "salida": format_dt(reg.fecha_salida),
        },
        "resumen": mensaje_asistencia,
        "mensaje": mensaje_full,
    })

@login_required(login_url='login')
def CambiarPasswordCliente(request):
    """Permite al cliente cambiar su contraseña"""
    if request.method == 'POST':
        password_actual = request.POST.get('password_actual')
        password_nueva = request.POST.get('password_nueva')
        password_confirmar = request.POST.get('password_confirmar')
        
        # Verificar contraseña actual
        if not request.user.check_password(password_actual):
            messages.error(request, 'La contraseña actual es incorrecta')
            return redirect('Perfil')
        
        # Verificar que las nuevas coincidan
        if password_nueva != password_confirmar:
            messages.error(request, 'Las contraseñas nuevas no coinciden')
            return redirect('Perfil')
        
        # Verificar longitud mínima
        if len(password_nueva) < 8:
            messages.error(request, 'La contraseña debe tener al menos 8 caracteres')
            return redirect('Perfil')
        
        # Cambiar contraseña
        request.user.set_password(password_nueva)
        request.user.save()
        
        # Mantener la sesión activa
        update_session_auth_hash(request, request.user)
        
        messages.success(request, '¡Contraseña actualizada exitosamente!')
        return redirect('Perfil')
    
    return redirect('Perfil')

def DetalleRutina(request, id):
    """Ver detalle de una rutina con sus ejercicios"""
    rutina = get_object_or_404(Rutina, Id_rutina=id, Estado='Activo')
    ejercicios = rutina.ejercicios.all().order_by('Orden')
    
    context = {
        'rutina': rutina,
        'ejercicios': ejercicios,
    }
    
    return render(request, 'templates_perfil/detalle_rutina.html', context)
