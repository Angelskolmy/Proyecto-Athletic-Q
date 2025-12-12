from django.shortcuts import render, redirect
from .models import producto 
from .models import categoria 
from .forms import ProductoForm  
from Historial.utils import registrar_movimiento
from django.contrib.auth.decorators import permission_required 
from django.contrib.auth.decorators import login_required 
from django.core.paginator import Paginator 
from django.http import HttpResponse
from openpyxl import Workbook 
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment  
from django.http import HttpResponse
from urllib.parse import urlencode 
from Detalle_venta.models import Detalle_Venta
from django.contrib import messages 

@permission_required('Productos.view_producto', raise_exception=True) 



def listarProductos (request): 

    AllProd= producto.objects.all().order_by('Id_producto')  

    paginator= Paginator(AllProd, 20)
    Page_number= request.GET.get('page') 
    Page_obj= paginator.get_page(Page_number)

    AllCatgo= categoria.objects.all() 
    Counter= producto.objects.all().count()

    ListProd= {'Productos' : Page_obj,
               'Categorias' : AllCatgo, 
               'Counter': Counter, 
                } 
    
    return render(request, 'templates_productos/productos.html', ListProd)


def IngresaProductos(request):  

    if request.method == 'POST': 
        Cipher= ProductoForm(request.POST, request.FILES)

        if Cipher.is_valid():
           
            ClaveHist= Cipher.save()  
            
            #----------  
            userList= request.user

            ListTIpo_Movimiento="ingresar" 
            LIstModulo="productos" 
            ListNombre_Objeto= ClaveHist.Nombre           
            ListId_Objeto= ClaveHist.Id_producto 
            Listid_usuario= userList 

            registrar_movimiento(
                user=Listid_usuario,
                tipo=ListTIpo_Movimiento,
                modulo=LIstModulo,
                nombre_objeto=ListNombre_Objeto,
                id_objeto=ListId_Objeto,
            )
            #----------
            return redirect('Producto')        
    else: 
        Cipher= ProductoForm()

    Clave={'Clave' : Cipher} 
    return render (request,'templates_productos/crear_productos.html', Clave)

@permission_required('Productos.delete_producto', raise_exception=True) 
def EliminarProducto(request, Id_producto):    

    Prueba= Detalle_Venta.objects.filter(Id_producto=Id_producto)

    if Prueba.exists():
        messages.error(request, 'El producto tiene ventas asosiadas no se puede borrar')  
        return redirect ('Producto')   
    
    else:   

        Shigaraki= producto.objects.get(Id_producto=Id_producto) 
    
        ListNombre_Objeto2= Shigaraki.Nombre
        ListId_Objeto2= Shigaraki.Id_producto

        Shigaraki.delete()  
        #------------------------
        UserLIst2= request.user 
        ListMdoulo2="productos" 
        ListTIpo_Movimiento2="eliminar"  
        registrar_movimiento(
            user=UserLIst2,
            tipo=ListTIpo_Movimiento2,
            modulo=ListMdoulo2,
            nombre_objeto=ListNombre_Objeto2,
            id_objeto=ListId_Objeto2,
        )
        return redirect ('Producto')  



def DetalleProducto (request, Id_producto): 

    consulta= " SELECT producto.*, categoria.nombre FROM producto join categoria on producto.Catego_Id = categoria.Id_categoria WHERE Id_producto= %s"
    EspecProd= producto.objects.raw(consulta, [Id_producto]) 
    ListEspec= { 'DetalleP' : EspecProd} 
    return render (request, 'templates_productos/detalle_productos.html' ,ListEspec)


def Editar_Producto (request, Id_producto): 

    ProdProt= producto.objects.get(Id_producto=Id_producto) 

    if request.method == 'POST': 

        Cifrado= ProductoForm(request.POST, request.FILES, instance=ProdProt) 

        if Cifrado.is_valid(): 
            
            Requiem= Cifrado.save()  
            
            ListUser3= request.user 
            ListNombre_Objeto3= Requiem.Nombre
            ListId_Objeto3=  Requiem.Id_producto
            ListMdoulo3= "productos"
            ListTIpo_Movimiento3= "editar"  
            registrar_movimiento(
                user=ListUser3,
                tipo=ListTIpo_Movimiento3,
                modulo=ListMdoulo3,
                nombre_objeto=ListNombre_Objeto3,
                id_objeto=ListId_Objeto3,
            )

            return redirect ('Producto') 
        
    else: 
        Cifrado= ProductoForm(instance=ProdProt) 
    
    Nomicon= {'Azat' : Cifrado} 

    return render (request, 'templates_productos/editar_productos.html' , Nomicon) 

def busqueda_producto (request): 

    catego_filter= request.GET.get('catego_filter') 
    Prodnombre_filter= request.GET.get('Prodnombre_filter', '').strip()  
    AllCategos2= categoria.objects.all() 

    if not Prodnombre_filter and catego_filter is None: 

        return redirect('Producto')

    if  catego_filter: 
        SQL_busqueda= "SELECT producto.*, categoria.nombre FROM producto join categoria on producto.Catego_Id = categoria.Id_categoria WHERE producto.Catego_Id= %s" 
        BusqPrd= producto.objects.raw(SQL_busqueda,[catego_filter]) 

    if Prodnombre_filter: 
        SQL_busqueda= "SELECT producto.*, categoria.nombre FROM producto join categoria on producto.Catego_Id = categoria.Id_categoria WHERE producto.Nombre= %s" 
        BusqPrd= producto.objects.raw(SQL_busqueda,[Prodnombre_filter]) 

    if catego_filter and Prodnombre_filter: 
        SQL_busqueda= "SELECT producto.*, categoria.nombre FROM producto join categoria on producto.Catego_Id = categoria.Id_categoria WHERE producto.Nombre= %s and producto.Catego_Id= %s" 
        BusqPrd= producto.objects.raw(SQL_busqueda,[Prodnombre_filter, catego_filter]) 

    Conversor= list(BusqPrd) 
    regulador= Paginator(Conversor,20) 
    Page_number= request.GET.get('page') 
    page_objt= regulador.get_page(Page_number)
     
    params = {k: v for k, v in request.GET.items() if k != 'page' and v != ''}
    querystring = '&' + urlencode(params) if params else ''


    Roku= {'niji' : page_objt,
           'Categorias': AllCategos2, 
           'querystring' : querystring} 
    return render (request, 'templates_productos/productos.html', Roku) 


def excel_content(request): 

    wb= Workbook() 
    ws = wb.active
    ws.title = "Productos"     

   # ======== ESTILOS ========

    # Fila superior grande (gris oscuro)
    top_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")

    # Encabezados (gris claro)
    header_fill = PatternFill(start_color="B7CDE8", end_color="B7CDE8", fill_type="solid")
    header_font = Font(bold=True, color="000000")

    # Bordes
    border = Border(
        left=Side(style='thin', color="000000"),
        right=Side(style='thin', color="000000"),
        top=Side(style='thin', color="000000"),
        bottom=Side(style='thin', color="000000")
    )

    # Alineación
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")


    # ======== PRIMERA FILA (barra gris grande) ========
    ws.merge_cells("A1:H1")
    cell_top = ws["A1"]
    cell_top.fill = top_fill
    cell_top.value = ""
    ws.row_dimensions[1].height = 30


    ws.append(['Id_producto','Catego_Id','Nombre','Descripcion','Stock','Precio_de_compra','Precio_de_venta','Estado']) 

    Alacran= producto.objects.all().order_by('Id_producto') 

    for rork in Alacran: 
        ws.append([rork.Id_producto, rork.Catego_Id.Nombre, rork.Nombre, rork.Descripcion, rork.Stock, rork.Precio_de_compra, rork.Precio_de_venta, rork.Estado])

    response= HttpResponse( 
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    ) 

    response['Content-Disposition'] = 'attachment; filename="Informe_productos.xlsx" ' 

    wb.save(response) 

    return (response)
