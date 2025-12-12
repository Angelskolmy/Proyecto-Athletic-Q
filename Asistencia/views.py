from io import BytesIO
from datetime import datetime, time

from django.contrib import messages
from django.contrib.auth.decorators import permission_required
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.enums import TA_CENTER
from reportlab.platypus import Paragraph, Spacer, Table, TableStyle, SimpleDocTemplate

from .forms import AsistenciaForm
from .models import Asistencia


@permission_required('Asistencia.view_asistencia', raise_exception=True)
def verPerfilasistencias(request):
    user = request.user
    if user.groups.filter(name='Admin').exists():
        return redirect('Asistencias')
    if user.groups.filter(name='Huella').exists():
        return render(request, "templates_asistencias/asistencias_huella.html")
    return redirect('Perfil')


def _parse_date(value: str):
    """Convierte 'YYYY-MM-DD' en date, o None si viene vacío/mal."""
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError:
        return None


def _day_bounds(date_obj):
    """
    Devuelve (inicio_de_día, fin_de_día) como datetimes *aware* si hay timezone.
    Ej: 2025-12-03 -> (2025-12-03 00:00, 2025-12-03 23:59:59.999999)
    """
    if not date_obj:
        return None, None

    start = datetime.combine(date_obj, time.min)
    end = datetime.combine(date_obj, time.max)

    # Si usas USE_TZ = True, hacemos los datetimes aware
    if timezone.is_naive(start):
        start = timezone.make_aware(start)
        end = timezone.make_aware(end)

    return start, end


def _filtered_asistencias_queryset(request):
    """
    Devuelve el queryset filtrado y los valores de filtros actuales para
    reutilizarlos tanto en la vista como en las exportaciones.
    """
    nombre_query = request.GET.get('nombre', '').strip()
    tipo_usuario = request.GET.get('tipo', '').strip()
    fecha_entrada_str = request.GET.get('fecha_entrada', '').strip()
    fecha_salida_str = request.GET.get('fecha_salida', '').strip()

    fecha_entrada = _parse_date(fecha_entrada_str)
    fecha_salida = _parse_date(fecha_salida_str)

    asistencias = Asistencia.objects.select_related('id_usuario').order_by('-fecha_entrada', '-id')

    if nombre_query:
        asistencias = asistencias.filter(
            Q(id_usuario__first_name__icontains=nombre_query) |
            Q(id_usuario__last_name__icontains=nombre_query) |
            Q(id_usuario__username__icontains=nombre_query) |
            Q(id_usuario__Cedula__icontains=nombre_query)
        )

    if tipo_usuario:
        tipo = tipo_usuario.lower()
        if tipo == 'admin':
            asistencias = asistencias.filter(
                Q(rol__iexact='Admin') |
                Q(rol__iexact='Administrador') |
                Q(rol__icontains='admin')
            )
        elif tipo == 'empleado':
            asistencias = asistencias.filter(
                Q(rol__iexact='Empleado') |
                Q(rol__iexact='Empleados') |
                Q(rol__icontains='empleado')
            )
        elif tipo == 'cliente':
            asistencias = asistencias.filter(
                Q(rol__iexact='Cliente') |
                Q(rol__iexact='Usuario') |
                Q(rol__iexact='Usuarios') |
                Q(rol__icontains='cliente') |
                Q(rol__icontains='usuario')
            )
        elif tipo == 'entrenador':
            asistencias = asistencias.filter(
                Q(rol__iexact='Entrenador') |
                Q(rol__icontains='entrenador')
            )
        else:
            asistencias = asistencias.filter(rol__icontains=tipo_usuario)

    if fecha_entrada and not fecha_salida:
        start, end = _day_bounds(fecha_entrada)
        asistencias = asistencias.filter(
            Q(fecha_entrada__range=(start, end)) |
            Q(fecha_salida__range=(start, end))
        )
    elif fecha_salida and not fecha_entrada:
        start, end = _day_bounds(fecha_salida)
        asistencias = asistencias.filter(
            Q(fecha_entrada__range=(start, end)) |
            Q(fecha_salida__range=(start, end))
        )
    elif fecha_entrada and fecha_salida:
        if fecha_entrada > fecha_salida:
            fecha_entrada, fecha_salida = fecha_salida, fecha_entrada

        start, _ = _day_bounds(fecha_entrada)
        _, end = _day_bounds(fecha_salida)
        asistencias = asistencias.filter(
            Q(fecha_entrada__range=(start, end)) |
            Q(fecha_salida__range=(start, end))
        )

    filtros = {
        "nombre_query": nombre_query,
        "tipo_usuario": tipo_usuario,
        "fecha_entrada": fecha_entrada_str,
        "fecha_salida": fecha_salida_str,
        "hay_filtros": any([nombre_query, tipo_usuario, fecha_entrada_str, fecha_salida_str]),
    }
    return asistencias, filtros


def _annotate_asistencias(queryset):
    """Calcula horas trabajadas y banderas auxiliares para cada registro."""
    asistencias_lista = list(queryset)
    for asistencia in asistencias_lista:
        cumplio_horario = None
        horas_trabajadas = None

        if asistencia.fecha_salida and asistencia.fecha_entrada:
            diferencia = asistencia.fecha_salida - asistencia.fecha_entrada
            horas_trabajadas = diferencia.total_seconds() / 3600
            cumplio_horario = horas_trabajadas >= 6

        asistencia.cumplio_horario = cumplio_horario
        asistencia.horas_trabajadas = round(horas_trabajadas, 2) if horas_trabajadas else None

    return asistencias_lista


@permission_required('Asistencia.view_asistencia', raise_exception=True)
def listarAsistencias(request):
    if request.user.groups.filter(name='Huella').exists():
        messages.error(request, 'No tienes acceso a esta vista')
        return redirect('AsisVista')
    page_number = request.GET.get('page', 1)

    asistencias_queryset, filtros = _filtered_asistencias_queryset(request)
    asistencias_lista = _annotate_asistencias(asistencias_queryset)

    paginator = Paginator(asistencias_lista, 15)
    page_obj = paginator.get_page(page_number)

    context = {
        "Asistencias": page_obj,
        "total_items": paginator.count,
    }
    context.update(filtros)
    
    return render(request, "templates_asistencias/asistencias.html", context)


@permission_required('Asistencia.view_asistencia', raise_exception=True)
def exportar_asistencias_excel(request):
    if request.user.groups.filter(name='Huella').exists():
        messages.error(request, 'No tienes acceso a esta vista')
        return redirect('AsisVista')

    asistencias_queryset, _ = _filtered_asistencias_queryset(request)
    asistencias_lista = _annotate_asistencias(asistencias_queryset)
    if not asistencias_lista:
        messages.error(request, "No hay asistencias para exportar.")
        return redirect('Asistencias')

    total = len(asistencias_lista)
    cumplidas = sum(1 for a in asistencias_lista if a.cumplio_horario is True)
    no_cumplidas = sum(1 for a in asistencias_lista if a.cumplio_horario is False)
    pendientes = total - (cumplidas + no_cumplidas)

    wb = Workbook()
    ws = wb.active
    ws.title = "Asistencias"

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 19
    ws.column_dimensions["E"].width = 19
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 16

    ws.merge_cells("A1:H1")
    ws["A1"] = "ATHLETIC-Q GIMNASIO"
    ws["A1"].font = Font(size=18, bold=True, color="8B0000")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:H2")
    ws["A2"] = "Reporte de Asistencias"
    ws["A2"].font = Font(size=13, bold=True)
    ws["A2"].alignment = Alignment(horizontal="center")

    info_rows = [
        ("Generado por:", request.user.get_full_name() or request.user.username, "Fecha:", timezone.now().strftime('%d/%m/%Y %H:%M')),
        ("Asistencias incluidas:", total, "Cumplieron horario:", cumplidas),
        ("No cumplieron:", no_cumplidas, "Pendientes:", pendientes),
    ]

    start_row = 4
    for left_label, left_val, right_label, right_val in info_rows:
        ws[f"A{start_row}"] = left_label
        ws[f"A{start_row}"].font = Font(bold=True)
        ws[f"B{start_row}"] = left_val
        ws[f"D{start_row}"] = right_label
        ws[f"D{start_row}"].font = Font(bold=True)
        ws[f"E{start_row}"] = right_val
        start_row += 1

    header_row = start_row + 1
    headers = ["#", "Usuario", "Rol", "Entrada", "Salida", "Horas", "Cumplió", "Estado"]
    header_style = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="8B0000")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col)
        cell.value = header
        cell.font = header_style
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    thin_border = Border(
        left=Side(style="thin", color="5A5A5A"),
        right=Side(style="thin", color="5A5A5A"),
        top=Side(style="thin", color="5A5A5A"),
        bottom=Side(style="thin", color="5A5A5A"),
    )

    current_row = header_row + 1
    for idx, asistencia in enumerate(asistencias_lista, 1):
        usuario = ""
        if asistencia.id_usuario:
            usuario = f"{asistencia.id_usuario.first_name} {asistencia.id_usuario.last_name}".strip() or asistencia.id_usuario.username
        cumplimiento = "Sí" if asistencia.cumplio_horario is True else "No" if asistencia.cumplio_horario is False else "Sin salida"
        values = [
            idx,
            usuario or "N/A",
            asistencia.rol or "—",
            asistencia.fecha_entrada.strftime("%d/%m/%Y %H:%M") if asistencia.fecha_entrada else "",
            asistencia.fecha_salida.strftime("%d/%m/%Y %H:%M") if asistencia.fecha_salida else "",
            asistencia.horas_trabajadas if asistencia.horas_trabajadas is not None else "",
            cumplimiento,
            asistencia.estado or "Sin estado",
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=current_row, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center") if col in (1, 6, 7, 8) else Alignment(vertical="center")
        current_row += 1

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=asistencias.xlsx'
    return response


@permission_required('Asistencia.view_asistencia', raise_exception=True)
def exportar_asistencias_pdf(request):
    if request.user.groups.filter(name='Huella').exists():
        messages.error(request, 'No tienes acceso a esta vista')
        return redirect('AsisVista')

    asistencias_queryset, _ = _filtered_asistencias_queryset(request)
    asistencias_lista = _annotate_asistencias(asistencias_queryset)
    if not asistencias_lista:
        messages.error(request, "No hay asistencias para exportar.")
        return redirect('Asistencias')

    total = len(asistencias_lista)
    cumplidas = sum(1 for a in asistencias_lista if a.cumplio_horario is True)
    no_cumplidas = sum(1 for a in asistencias_lista if a.cumplio_horario is False)
    pendientes = total - (cumplidas + no_cumplidas)

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'AsistenciaTitle',
        parent=styles['Heading1'],
        fontSize=20,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#8B0000')
    )
    subtitle_style = ParagraphStyle(
        'AsistenciaSubtitle',
        parent=styles['Heading2'],
        fontSize=13,
        alignment=TA_CENTER,
        spaceAfter=10,
    )
    normal_style = ParagraphStyle(
        "AsistenciaNormal",
        parent=styles["Normal"],
        fontSize=10,
        spaceAfter=4,
    )

    elements = [
        Paragraph("ATHLETIC-Q GIMNASIO", title_style),
        Paragraph("Reporte de Asistencias", subtitle_style),
        Paragraph(
            f"<b>Generado por:</b> {request.user.get_full_name() or request.user.username}",
            normal_style,
        ),
        Paragraph(
            f"<b>Fecha:</b> {timezone.now().strftime('%d/%m/%Y %H:%M')}",
            normal_style,
        ),
        Spacer(1, 12)
    ]

    data = [["#", "Usuario", "Rol", "Entrada", "Salida", "Horas", "Cumplió", "Estado"]]
    for idx, asistencia in enumerate(asistencias_lista, 1):
        usuario = ""
        if asistencia.id_usuario:
            usuario = f"{asistencia.id_usuario.first_name} {asistencia.id_usuario.last_name}".strip()
            if not usuario:
                usuario = asistencia.id_usuario.username
        fecha_entrada = asistencia.fecha_entrada.strftime("%d/%m/%Y %H:%M") if asistencia.fecha_entrada else "—"
        fecha_salida = asistencia.fecha_salida.strftime("%d/%m/%Y %H:%M") if asistencia.fecha_salida else "—"

        if asistencia.cumplio_horario is True:
            cumplimiento = "Sí"
        elif asistencia.cumplio_horario is False:
            cumplimiento = "No"
        else:
            cumplimiento = "Sin salida"

        data.append([
            idx,
            usuario or "N/A",
            asistencia.rol or "—",
            fecha_entrada,
            fecha_salida,
            asistencia.horas_trabajadas if asistencia.horas_trabajadas is not None else "—",
            cumplimiento,
            asistencia.estado or "Sin estado",
        ])

    table = Table(data, repeatRows=1, colWidths=[25, 110, 60, 80, 80, 45, 55, 70])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#8B0000')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f7f7f7')]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))

    elements.append(table)
    elements.append(Spacer(1, 18))

    resumen = [
        ["Asistencias incluidas:", str(total)],
        ["Cumplieron horario:", str(cumplidas)],
        ["No cumplieron:", str(no_cumplidas)],
        ["Pendientes:", str(pendientes)],
    ]
    resumen_table = Table(resumen, colWidths=[170, 70])
    resumen_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#444444')),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(resumen_table)
    elements.append(Spacer(1, 20))

    footer_style = ParagraphStyle(
        'AsistenciaFooter',
        parent=styles['Normal'],
        fontSize=8,
        alignment=TA_CENTER,
        textColor=colors.grey,
    )
    elements.append(Paragraph("──────────────────────────────────────────────", footer_style))
    elements.append(Spacer(1, 6))
    elements.append(Paragraph("Reporte generado por Athletic-Q | Gestión de Asistencias", footer_style))

    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=asistencias.pdf'
    response.write(pdf)
    return response
