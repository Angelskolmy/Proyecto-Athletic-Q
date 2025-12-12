from io import BytesIO
from datetime import datetime, time
from decimal import Decimal

from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.core.paginator import Paginator
from django.db.models import Q, Sum, Count
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import Paragraph, Spacer, Table, TableStyle, SimpleDocTemplate

from .models import Historial_Ventas

def _parse_date(value: str):
    """Convierte 'YYYY-MM-DD' en date, o None si viene vacío/mal."""
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError:
        return None


def _day_bounds(date_obj):
    """
    Devuelve (inicio_de_día, fin_de_día) como datetimes *aware* si hay timezone.
    """
    if not date_obj:
        return None, None

    start = datetime.combine(date_obj, time.min)
    end = datetime.combine(date_obj, time.max)

    # Si usas USE_TZ = True, hacemos los datetimes aware
    if timezone.is_naive(start):
        start = timezone.make_aware(start)
        end = timezone.make_aware(end)

    return start, end


def _filtered_historial_queryset(request):
    search_query = request.GET.get('search', '').strip()
    filter_metodo = request.GET.get('metodo', '').strip()
    filter_fecha_str = request.GET.get('fecha', '').strip()
    filter_fecha = _parse_date(filter_fecha_str)
    historial = Historial_Ventas.objects.select_related(
        'id_usuario', 
        'id_venta'
    ).order_by('-id_registro')

    if search_query:
        historial = historial.filter(
            Q(id_usuario__first_name__icontains=search_query) |
            Q(id_usuario__last_name__icontains=search_query) |
            Q(id_usuario__Cedula__icontains=search_query) |
            Q(id_venta__Cedula_Vents__icontains=search_query) |
            Q(id_registro__icontains=search_query) |
            Q(id_venta__Id_venta__icontains=search_query)
        )

    if filter_metodo:
        historial = historial.filter(metodo_pago__iexact=filter_metodo)

    if filter_fecha:
        start, end = _day_bounds(filter_fecha)
        if start and end:
            historial = historial.filter(fecha_venta__range=(start, end))

    filtros = {
        'search_query': search_query,
        'filter_metodo': filter_metodo,
        'filter_fecha': filter_fecha_str,
        'hay_filtros': any([search_query, filter_metodo, filter_fecha_str]),
    }
    return historial, filtros


@permission_required('Historial_ventas.view_historial_ventas', raise_exception=True)
def ListarHistorialVentas(request):
    historial, filtros = _filtered_historial_queryset(request)
    page_number = request.GET.get('page', 1)

    stats = historial.aggregate(
        total_ventas=Count('id_registro'),
        total_recaudado=Sum('Monto')
    )
    
    total_ventas = stats['total_ventas'] or 0
    total_recaudado = stats['total_recaudado'] or Decimal('0.00')

    paginator = Paginator(historial, 10)
    page_obj = paginator.get_page(page_number)

    context = {
        'AllHV': page_obj,
        'total_items': paginator.count,
        'total_ventas': total_ventas,
        'total_recaudado': total_recaudado,
    }
    context.update(filtros)

    return render(request, 'templates_ventas/historial_ventas.html', context)


@login_required(login_url='login')
@permission_required('Historial_ventas.view_historial_ventas', raise_exception=True)
def exportar_historial_ventas_excel(request):
    historial, filtros = _filtered_historial_queryset(request)
    registros = list(historial)
    if not registros:
        messages.error(request, "No hay registros de ventas para exportar.")
        return redirect('HistorialVentas')

    stats = historial.aggregate(
        total_ventas=Count('id_registro'),
        total_recaudado=Sum('Monto')
    )
    total_ventas = stats['total_ventas'] or 0
    total_recaudado = stats['total_recaudado'] or Decimal('0.00')

    resumen_metodos = historial.values('metodo_pago').annotate(
        total=Count('id_registro'),
        monto=Sum('Monto')
    ).order_by('metodo_pago')

    wb = Workbook()
    ws = wb.active
    ws.title = "Historial Ventas"

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 20
    ws.column_dimensions["H"].width = 16

    ws.merge_cells("A1:H1")
    ws["A1"] = "ATHLETIC-Q GIMNASIO"
    ws["A1"].font = Font(size=18, bold=True, color="8B0000")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:H2")
    ws["A2"] = "Historial de Ventas"
    ws["A2"].font = Font(size=13, bold=True)
    ws["A2"].alignment = Alignment(horizontal="center")

    info_rows = [
        ("Generado por:", request.user.get_full_name() or request.user.username, "Fecha:", timezone.now().strftime('%d/%m/%Y %H:%M')),
        ("Ventas incluidas:", total_ventas, "Total recaudado:", f"${total_recaudado:,.0f}"),
    ]
    start_row = 4
    for left_label, left_value, right_label, right_value in info_rows:
        ws[f"A{start_row}"] = left_label
        ws[f"A{start_row}"].font = Font(bold=True)
        ws[f"B{start_row}"] = left_value
        ws[f"D{start_row}"] = right_label
        ws[f"D{start_row}"].font = Font(bold=True)
        ws[f"E{start_row}"] = right_value
        start_row += 1

    header_row = start_row + 1
    headers = ["#", "ID Venta", "Cédula Cliente", "Cédula Empleado", "Empleado", "Método de Pago", "Fecha", "Total"]
    header_style = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="8B0000")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col)
        cell.value = header
        cell.font = header_style
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    thin_border = Border(
        left=Side(style="thin", color="5A5A5A"),
        right=Side(style="thin", color="5A5A5A"),
        top=Side(style="thin", color="5A5A5A"),
        bottom=Side(style="thin", color="5A5A5A"),
    )

    current_row = header_row + 1
    for idx, registro in enumerate(registros, 1):
        cliente_cedula = registro.id_venta.Cedula_Vents or "—"
        empleado_cedula = registro.id_usuario.Cedula if registro.id_usuario else "—"
        empleado_nombre = ""
        if registro.id_usuario:
            empleado_nombre = f"{registro.id_usuario.first_name} {registro.id_usuario.last_name}".strip()
        fecha = registro.fecha_venta.strftime("%d/%m/%Y %H:%M") if registro.fecha_venta else ""

        values = [
            idx,
            registro.id_venta.Id_venta if registro.id_venta else "—",
            cliente_cedula,
            empleado_cedula,
            empleado_nombre or "—",
            registro.metodo_pago or "—",
            fecha,
            float(registro.Monto),
        ]

        for col, value in enumerate(values, 1):
            cell = ws.cell(row=current_row, column=col, value=value)
            cell.border = thin_border
            if col in (1, 2, 3, 4, 6):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(vertical="center")
            if col == 8:
                cell.number_format = "$#,##0"

        current_row += 1

    summary_start = current_row + 1
    ws.cell(row=summary_start, column=6, value="Resumen por método").font = Font(bold=True)
    summary_row = summary_start + 1
    for item in resumen_metodos:
        ws.cell(row=summary_row, column=6, value=item['metodo_pago'])
        ws.cell(row=summary_row, column=7, value=item['total'])
        monto_cell = ws.cell(row=summary_row, column=8, value=float(item['monto'] or 0))
        monto_cell.number_format = "$#,##0"
        summary_row += 1

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=historial_ventas.xlsx'
    return response


@login_required(login_url='login')
@permission_required('Historial_ventas.view_historial_ventas', raise_exception=True)
def exportar_historial_ventas_pdf(request):
    historial, filtros = _filtered_historial_queryset(request)
    registros = list(historial)
    if not registros:
        messages.error(request, "No hay registros de ventas para exportar.")
        return redirect('HistorialVentas')

    stats = historial.aggregate(
        total_ventas=Count('id_registro'),
        total_recaudado=Sum('Monto')
    )
    total_ventas = stats['total_ventas'] or 0
    total_recaudado = stats['total_recaudado'] or Decimal('0.00')

    resumen_metodos = historial.values('metodo_pago').annotate(
        total=Count('id_registro'),
        monto=Sum('Monto')
    ).order_by('metodo_pago')

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'VentasTitle',
        parent=styles['Heading1'],
        fontSize=20,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#8B0000')
    )
    subtitle_style = ParagraphStyle(
        'VentasSubtitle',
        parent=styles['Heading2'],
        fontSize=13,
        alignment=TA_CENTER,
        spaceAfter=10,
    )
    normal_style = ParagraphStyle(
        'VentasNormal',
        parent=styles['Normal'],
        fontSize=10,
        spaceAfter=4,
    )

    elements = [
        Paragraph("ATHLETIC-Q GIMNASIO", title_style),
        Paragraph("Historial de Ventas", subtitle_style),
        Paragraph(
            f"<b>Generado por:</b> {request.user.get_full_name() or request.user.username}",
            normal_style,
        ),
        Paragraph(
            f"<b>Fecha:</b> {timezone.now().strftime('%d/%m/%Y %H:%M')}",
            normal_style,
        ),
        Spacer(1, 12)
    ]

    data = [["#", "ID Venta", "Céd. Cliente", "Céd. Empleado", "Empleado", "Método", "Fecha", "Total"]]
    for idx, registro in enumerate(registros, 1):
        cliente_cedula = registro.id_venta.Cedula_Vents or "—"
        empleado_cedula = registro.id_usuario.Cedula if registro.id_usuario else "—"
        empleado_nombre = ""
        if registro.id_usuario:
            empleado_nombre = f"{registro.id_usuario.first_name} {registro.id_usuario.last_name}".strip()
        fecha = registro.fecha_venta.strftime("%d/%m/%Y %H:%M") if registro.fecha_venta else "—"

        data.append([
            idx,
            registro.id_venta.Id_venta if registro.id_venta else "—",
            cliente_cedula,
            empleado_cedula,
            empleado_nombre or "—",
            registro.metodo_pago or "—",
            fecha,
            f"${registro.Monto:,.0f}",
        ])

    table = Table(data, repeatRows=1, colWidths=[25, 60, 70, 70, 100, 70, 80, 60])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#8B0000')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f7f7f7')]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))

    elements.append(table)
    elements.append(Spacer(1, 18))

    resumen_data = [
        ["Ventas incluidas:", str(total_ventas)],
        ["Total recaudado:", f"${total_recaudado:,.0f}"],
    ]
    resumen_data.append(["", ""])
    for item in resumen_metodos:
        resumen_data.append([
            f"{item['metodo_pago']}:", f"{item['total']} ventas | ${item['monto'] or 0:,.0f}"
        ])

    resumen_table = Table(resumen_data, colWidths=[160, 140])
    resumen_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#444444')),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(resumen_table)
    elements.append(Spacer(1, 20))

    footer_style = ParagraphStyle(
        'VentasFooter',
        parent=styles['Normal'],
        fontSize=8,
        alignment=TA_CENTER,
        textColor=colors.grey,
    )
    elements.append(Paragraph("──────────────────────────────────────────────", footer_style))
    elements.append(Spacer(1, 6))
    elements.append(Paragraph("Reporte generado por Athletic-Q | Historial de Ventas", footer_style))

    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=historial_ventas.pdf'
    response.write(pdf)
    return response
