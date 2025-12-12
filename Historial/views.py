from io import BytesIO

from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.core.paginator import Paginator
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.utils.dateparse import parse_date
from django.utils import timezone
from urllib.parse import urlencode

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from reportlab.platypus import Paragraph, Spacer, Table, TableStyle, SimpleDocTemplate

from .models import Historial_usuario

def _build_historial_queryset(request):
    qs = Historial_usuario.objects.select_related('id_usuario').order_by('-Id_historial')

    hist_emp = request.GET.get('HistEmp', '').strip()
    hist_mov = request.GET.get('HistMovimientos', '').strip()
    hist_mod = request.GET.get('HistModulo', '').strip()
    hist_fecha = request.GET.get('HistFecha', '').strip()

    if hist_emp:
        qs = qs.filter(id_usuario__first_name__icontains=hist_emp)

    if hist_mov:
        qs = qs.filter(TIpo_Movimiento__iexact=hist_mov)

    if hist_mod:
        qs = qs.filter(Modulo__iexact=hist_mod)

    if hist_fecha:
        fecha = parse_date(hist_fecha)
        if fecha:
            qs = qs.filter(Fecha_y_hora=fecha)

    return qs


def listHistU(request):
    queryset = _build_historial_queryset(request)
    paginator = Paginator(queryset, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    params = {k: v for k, v in request.GET.items() if k != 'page' and v}
    querystring = '&' + urlencode(params) if params else ''

    context = {
        'HistU': page_obj,
        'Filter': None,
        'Conter': paginator.count,
        'querystring': querystring,
    }
    return render(request, 'templates_historial_usuario/historial_usuario.html', context)


def BusqHistorial(request):
    if not any(request.GET.get(k) for k in ['HistEmp', 'HistMovimientos', 'HistModulo', 'HistFecha']):
        return redirect('HistorialU')

    queryset = _build_historial_queryset(request)
    paginator = Paginator(queryset, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    params = {k: v for k, v in request.GET.items() if k != 'page' and v}
    querystring = '&' + urlencode(params) if params else ''

    context = {
        'HistU': None,
        'Filter': page_obj,
        'Conter': paginator.count,
        'querystring': querystring,
    }
    return render(request, 'templates_historial_usuario/historial_usuario.html', context)


@login_required(login_url='login')
@permission_required('Historial.view_historial_usuario', raise_exception=True)
def exportar_historial_excel(request):
    queryset = _build_historial_queryset(request)
    registros = list(queryset)
    if not registros:
        messages.error(request, "No hay movimientos para exportar.")
        return redirect('HistorialU')

    total = len(registros)
    movimientos_resumen = {"Ingresar": 0, "Editar": 0, "Eliminar": 0}
    for registro in registros:
        mov = (registro.TIpo_Movimiento or "").capitalize()
        if mov in movimientos_resumen:
            movimientos_resumen[mov] += 1

    wb = Workbook()
    ws = wb.active
    ws.title = "Historial"

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 28
    ws.column_dimensions["G"].width = 12

    ws.merge_cells("A1:G1")
    ws["A1"] = "ATHLETIC-Q GIMNASIO"
    ws["A1"].font = Font(size=18, bold=True, color="8B0000")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:G2")
    ws["A2"] = "Reporte de Movimientos"
    ws["A2"].font = Font(size=13, bold=True)
    ws["A2"].alignment = Alignment(horizontal="center")

    info_rows = [
        ("Generado por:", request.user.get_full_name() or request.user.username, "Fecha:", timezone.now().strftime('%d/%m/%Y %H:%M')),
        ("Movimientos incluidos:", total, "Ingresar:", movimientos_resumen["Ingresar"]),
        ("Editar:", movimientos_resumen["Editar"], "Eliminar:", movimientos_resumen["Eliminar"]),
    ]

    start_row = 4
    for left_label, left_value, right_label, right_value in info_rows:
        ws[f"A{start_row}"] = left_label
        ws[f"A{start_row}"].font = Font(bold=True)
        ws[f"B{start_row}"] = left_value
        ws[f"D{start_row}"] = right_label
        ws[f"D{start_row}"].font = Font(bold=True)
        ws[f"E{start_row}"] = right_value
        start_row += 1

    header_row = start_row + 1
    headers = ["#", "Fecha", "Usuario", "Movimiento", "Módulo", "Objeto", "ID Objeto"]
    header_style = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="8B0000")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col)
        cell.value = header
        cell.font = header_style
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    thin_border = Border(
        left=Side(style="thin", color="5A5A5A"),
        right=Side(style="thin", color="5A5A5A"),
        top=Side(style="thin", color="5A5A5A"),
        bottom=Side(style="thin", color="5A5A5A"),
    )

    current_row = header_row + 1
    for idx, registro in enumerate(registros, 1):
        values = [
            idx,
            registro.Fecha_y_hora.strftime("%d/%m/%Y %H:%M") if registro.Fecha_y_hora else "—",
            registro.id_usuario.get_full_name() if registro.id_usuario else "N/A",
            registro.TIpo_Movimiento or "—",
            registro.Modulo or "—",
            registro.Nombre_Objeto or "—",
            registro.Id_Objeto or "—",
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=current_row, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center") if col in (1, 4, 7) else Alignment(vertical="center")
        current_row += 1

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=historial_movimientos.xlsx'
    return response


@login_required(login_url='login')
@permission_required('Historial.view_historial_usuario', raise_exception=True)
def exportar_historial_pdf(request):
    queryset = _build_historial_queryset(request)
    registros = list(queryset)
    if not registros:
        messages.error(request, "No hay movimientos para exportar.")
        return redirect('HistorialU')

    total = len(registros)
    movimientos_resumen = {"Ingresar": 0, "Editar": 0, "Eliminar": 0}
    for registro in registros:
        mov = (registro.TIpo_Movimiento or "").capitalize()
        if mov in movimientos_resumen:
            movimientos_resumen[mov] += 1

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'Title',
        parent=styles['Heading1'],
        fontSize=20,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#8B0000')
    )
    subtitle_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Heading2'],
        fontSize=13,
        alignment=TA_CENTER,
        spaceAfter=10,
    )
    normal_style = ParagraphStyle(
        'NormalCustom',
        parent=styles['Normal'],
        fontSize=10,
        spaceAfter=4,
    )

    elements = [
        Paragraph("ATHLETIC-Q GIMNASIO", title_style),
        Paragraph("Reporte de Movimientos", subtitle_style),
        Paragraph(
            f"<b>Generado por:</b> {request.user.get_full_name() or request.user.username}",
            normal_style,
        ),
        Paragraph(
            f"<b>Fecha:</b> {timezone.now().strftime('%d/%m/%Y %H:%M')}",
            normal_style,
        ),
        Spacer(1, 12)
    ]

    data = [["#", "Fecha", "Usuario", "Movimiento", "Módulo", "Objeto", "ID Objeto"]]
    for idx, registro in enumerate(registros, 1):
        data.append([
            idx,
            registro.Fecha_y_hora.strftime("%d/%m/%Y %H:%M") if registro.Fecha_y_hora else "—",
            registro.id_usuario.get_full_name() if registro.id_usuario else "N/A",
            registro.TIpo_Movimiento or "—",
            registro.Modulo or "—",
            registro.Nombre_Objeto or "—",
            registro.Id_Objeto or "—",
        ])

    table = Table(data, repeatRows=1, colWidths=[25, 80, 120, 70, 80, 100, 60])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#8B0000')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f7f7f7')]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))

    elements.append(table)
    elements.append(Spacer(1, 18))

    resumen_data = [
        ["Movimientos incluidos:", str(total)],
        ["Ingresar:", str(movimientos_resumen["Ingresar"])],
        ["Editar:", str(movimientos_resumen["Editar"])],
        ["Eliminar:", str(movimientos_resumen["Eliminar"])],
    ]
    resumen_table = Table(resumen_data, colWidths=[160, 60])
    resumen_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#444444')),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))

    elements.append(resumen_table)
    elements.append(Spacer(1, 20))

    footer_style = ParagraphStyle(
        'HistorialFooter',
        parent=styles['Normal'],
        fontSize=8,
        alignment=TA_CENTER,
        textColor=colors.grey,
    )
    elements.append(Paragraph("──────────────────────────────────────────────", footer_style))
    elements.append(Spacer(1, 6))
    elements.append(Paragraph("Reporte generado por Athletic-Q | Gestión de Movimientos", footer_style))

    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=historial_movimientos.pdf'
    response.write(pdf)
    return response
