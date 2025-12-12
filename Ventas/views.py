from django.shortcuts import render, redirect, get_object_or_404
from django.db import transaction
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib import messages
from django.db.models import Sum, Count
from django.utils import timezone
from django.http import JsonResponse, HttpResponse
from decimal import Decimal
from io import BytesIO

# Para Excel
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# Para PDF
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

from .models import Venta
from .forms import VentaForm
from Detalle_venta.models import Detalle_Venta
from Productos.models import producto
from Empleados.models import User_Empleados
from Categorias.models import categoria
from Historial_ventas.models import Historial_Ventas 
from Historial.utils import registrar_movimiento


@login_required(login_url='login')
@permission_required('Ventas.add_venta', login_url='home')
def ListarVentas(request):
    """Vista principal de ventas con estadísticas del empleado"""
    
    empleado = request.user
    hoy = timezone.now().date()
    
    # Obtener ventas del empleado en el día
    ventas_hoy = Venta.objects.filter(
        id_usuario=empleado,
        Fecha__date=hoy
    )
    
    # Calcular estadísticas
    stats = ventas_hoy.aggregate(
        total_ventas=Count('Id_venta'),
        total_recaudado=Sum('Total')
    )
    
    total_ventas = stats['total_ventas'] or 0
    total_recaudado = stats['total_recaudado'] or Decimal('0.00')
    
    # Datos para formulario
    empleados = User_Empleados.objects.filter(is_active=True, is_staff=True).order_by('first_name')
    productos = producto.objects.filter(Estado='Activo').select_related('Catego_Id').order_by('Nombre')
    categorias = categoria.objects.filter(Estado='Activo').order_by('Nombre')
    
    # Obtener clientes registrados (usuarios del grupo Usuarios)
    clientes = User_Empleados.objects.filter(
        groups__name='Usuarios',
        is_active=True
    ).order_by('first_name')
    
    context = {
        'empleado': empleado,
        'total_ventas': total_ventas,
        'total_recaudado': total_recaudado,
        'empleados': empleados,
        'productos': productos,
        'categorias': categorias,
        'clientes': clientes, 
        'hoy': hoy,
    }
    
    return render(request, "templates_ventas/ventas.html", context)

@login_required(login_url='login')
@permission_required('Ventas.add_venta', login_url='Ventas')
@transaction.atomic
def ProcesarVenta(request):
    """Procesar nueva venta"""
    
    if request.method != 'POST':
        messages.error(request, 'Método no permitido')
        return redirect('Ventas')
    
    try:
        # Obtener vendedor
        empleado_id = request.POST.get('empleado_id')
        
        # Validar que empleado_id no esté vacío
        if not empleado_id:
            messages.error(request, 'Debe seleccionar un vendedor')
            return redirect('Ventas')
        
        empleado = get_object_or_404(User_Empleados, id=empleado_id)
        
        # 2. Obtener cliente (registrado o fantasma)
        cliente_id = request.POST.get('cliente_id')
        
        if not cliente_id:
            messages.error(request, 'Debe seleccionar un cliente')
            return redirect('Ventas')
        
        cliente = get_object_or_404(User_Empleados, id=cliente_id)
        cedula_vents = cliente.Cedula  # Tomar cédula del cliente seleccionado
        
        # Obtener metodo de pago
        metodo_pago = request.POST.get('metodo_pago', 'Efectivo')
        
        # Obtener productos
        productos_ids = request.POST.getlist('producto_id[]')
        cantidades = request.POST.getlist('cantidad[]')
        
        if not productos_ids:
            messages.error(request, 'Debe agregar al menos un producto')
            return redirect('Ventas')
        
        # Calcular total
        total = Decimal('0.00')
        items = []
        
        for i in range(len(productos_ids)):
            prod = get_object_or_404(producto, Id_producto=productos_ids[i])
            cantidad = int(cantidades[i])
            
            if prod.Stock < cantidad:
                messages.error(request, f'Stock insuficiente para {prod.Nombre}')
                return redirect('Ventas')
            
            subtotal = prod.Precio_de_venta * cantidad
            total += subtotal
            
            items.append({
                'producto': prod,
                'cantidad': cantidad,
                'precio': prod.Precio_de_venta,
                'subtotal': subtotal
            })
        
        # Crear venta
        venta = Venta.objects.create(
            id_usuario=empleado,
            Total=total,
            Cedula_Vents=cedula_vents
        )
        
        # Crear detalles y descontar stock
        for item in items:
            item['producto'].Stock -= item['cantidad']
            item['producto'].save()
            
            Detalle_Venta.objects.create(
                Id_venta=venta,
                Id_producto=item['producto'],
                Tipo_Pago=metodo_pago,
                Cantidad=item['cantidad'],
                Subtotal=item['subtotal'],
                Total=total
            )
        
        # Registrar en el historial 
        Historial_Ventas.objects.create(
            id_usuario=empleado,
            id_venta=venta,
            Monto=total,
            metodo_pago=metodo_pago
        ) 

        registrar_movimiento(
            user=empleado,
            tipo='ingresar',
            modulo='ventas',
            nombre_objeto=f"Venta #{venta.Id_venta}",
            id_objeto=venta.Id_venta,
        )

        messages.success(request, f' Venta #{venta.Id_venta} creada - Total: ${total:,.0f}')
        return redirect('Ventas')
        
    except Exception as e:
        messages.error(request, f' Error: {str(e)}')
        return redirect('Ventas')


@login_required(login_url='login')
@permission_required('Ventas.view_venta', login_url='Ventas')
def DetalleVenta(request, id):
    """Ver detalle de una venta"""
    
    venta = get_object_or_404(
        Venta.objects.select_related('id_usuario'),
        Id_venta=id
    )
    
    detalles = Detalle_Venta.objects.filter(
        Id_venta=venta
    ).select_related('Id_producto__Catego_Id')
    
    total_items = sum(detalle.Cantidad for detalle in detalles)
    subtotal_sin_iva = venta.Total / Decimal('1.19')
    iva = venta.Total - subtotal_sin_iva
    
    context = {
        'venta': venta,
        'detalles': detalles,
        'total_items': total_items,
        'subtotal_sin_iva': subtotal_sin_iva,
        'iva': iva,
    }
    
    return render(request, 'templates_ventas/detalle_venta.html', context)


@login_required(login_url='login')
@permission_required('Ventas.change_venta', login_url='Ventas')
def EditarVenta(request, id):
    """Editar venta - SOLO ADMIN"""
    
    venta = get_object_or_404(Venta, Id_venta=id)
    detalles = Detalle_Venta.objects.filter(Id_venta=venta).select_related('Id_producto__Catego_Id')
    
    empleados = User_Empleados.objects.filter(is_active=True, is_staff=True).order_by('first_name')
    productos = producto.objects.filter(Estado='Activo').select_related('Catego_Id').order_by('Nombre')
    categorias = categoria.objects.filter(Estado='Activo').order_by('Nombre')
    
    # Preparar carrito con los productos actuales
    carrito_inicial = []
    for detalle in detalles:
        carrito_inicial.append({
            'id': detalle.Id_producto.Id_producto,
            'nombre': detalle.Id_producto.Nombre,
            'precio': float(detalle.Id_producto.Precio_de_venta),
            'cantidad': detalle.Cantidad,
            'stock': detalle.Id_producto.Stock + detalle.Cantidad
        })
    
    return render(request, 'templates_ventas/editar_ventas.html', {
        'venta': venta,
        'detalles': detalles,
        'empleados': empleados,
        'productos': productos,
        'categorias': categorias,
        'carrito_inicial': carrito_inicial,
    })


@login_required(login_url='login')
@permission_required('Ventas.change_venta', login_url='Ventas')
@transaction.atomic
def ActualizarVenta(request, id):
    """Actualizar venta existente"""
    
    if request.method != 'POST':
        messages.error(request, 'Método no permitido')
        return redirect('Ventas')
    
    try:
        venta = get_object_or_404(Venta, Id_venta=id)
        detalles_anteriores = Detalle_Venta.objects.filter(Id_venta=venta)
        
        # Restaurar stock de productos anteriores
        for detalle in detalles_anteriores:
            detalle.Id_producto.Stock += detalle.Cantidad
            detalle.Id_producto.save()
        
        # Eliminar detalles anteriores
        detalles_anteriores.delete()
        
        # Obtener nuevos datos
        metodo_pago = request.POST.get('metodo_pago', 'Efectivo')
        observaciones = request.POST.get('observaciones', '')
        
        productos_ids = request.POST.getlist('producto_id[]')
        cantidades = request.POST.getlist('cantidad[]')
        
        if not productos_ids:
            messages.error(request, 'Debe agregar al menos un producto')
            return redirect('editar_venta', id=id)
        
        # Calcular nuevo total
        total = Decimal('0.00')
        
        for i in range(len(productos_ids)):
            prod = get_object_or_404(producto, Id_producto=productos_ids[i])
            cantidad = int(cantidades[i])
            
            if prod.Stock < cantidad:
                messages.error(request, f'Stock insuficiente para {prod.Nombre}')
                return redirect('editar_venta', id=id)
            
            subtotal = prod.Precio_de_venta * cantidad
            total += subtotal
            
            # Descontar nuevo stock
            prod.Stock -= cantidad
            prod.save()
            
            # Crear nuevo detalle
            Detalle_Venta.objects.create(
                Id_venta=venta,
                Id_producto=prod,
                Tipo_Pago=metodo_pago,
                Cantidad=cantidad,
                Subtotal=subtotal,
                Total=total
            )
        
        # Actualizar venta
        venta.Total = total
        venta.observaciones_edicion = observaciones
        venta.save() 

        histVend2 = request.user
        histMod2 = 'ventas'
        registrar_movimiento(
            user=histVend2,
            tipo='editar',
            modulo='ventas',
            nombre_objeto=f"Venta #{venta.Id_venta}",
            id_objeto=venta.Id_venta,
        )

        
        messages.success(request, f' Venta #{venta.Id_venta} actualizada exitosamente')
        return redirect('detalle_venta', id=venta.Id_venta)
        
    except Exception as e:
        messages.error(request, f' Error: {str(e)}')
        return redirect('editar_venta', id=id)
    
# =============================================
# EXPORTAR A EXCEL
# =============================================
@login_required(login_url='login')
@permission_required('Ventas.view_venta', login_url='Ventas')
def ExportarVentaExcel(request, id):
    """Exportar detalle de venta a Excel"""
    
    venta = get_object_or_404(Venta.objects.select_related('id_usuario'), Id_venta=id)
    detalles = Detalle_Venta.objects.filter(Id_venta=venta).select_related('Id_producto__Catego_Id')
    
    # Crear libro de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = f"Venta_{venta.Id_venta}"
    
    # Estilos
    header_fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14, color="8B0000")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    
    # ========== ENCABEZADO DE LA EMPRESA ==========
    ws.merge_cells('A1:F1')
    ws['A1'] = "ATHLETIC-Q GIMNASIO"
    ws['A1'].font = Font(bold=True, size=18, color="8B0000")
    ws['A1'].alignment = center
    
    ws.merge_cells('A2:F2')
    ws['A2'] = "Comprobante de Venta"
    ws['A2'].font = Font(bold=True, size=14)
    ws['A2'].alignment = center
    
    # ========== INFORMACIÓN DE LA VENTA ==========
    ws['A4'] = "Venta #:"
    ws['A4'].font = Font(bold=True)
    ws['B4'] = venta.Id_venta
    
    ws['D4'] = "Fecha:"
    ws['D4'].font = Font(bold=True)
    ws['E4'] = venta.Fecha.strftime('%d/%m/%Y %H:%M')
    
    ws['A5'] = "Vendedor:"
    ws['A5'].font = Font(bold=True)
    ws['B5'] = f"{venta.id_usuario.first_name} {venta.id_usuario.last_name}" if venta.id_usuario else "N/A"
    
    ws['D5'] = "Método de Pago:"
    ws['D5'].font = Font(bold=True)
    primer_detalle = detalles.first()
    ws['E5'] = primer_detalle.Tipo_Pago if primer_detalle else "N/A"
    
    if venta.Cedula_Vents:
        ws['A6'] = "Cédula Cliente:"
        ws['A6'].font = Font(bold=True)
        ws['B6'] = venta.Cedula_Vents
    
    if venta.Numero_Transaccion:
        ws['D6'] = "# Transacción:"
        ws['D6'].font = Font(bold=True)
        ws['E6'] = venta.Numero_Transaccion
    
    # ========== TABLA DE PRODUCTOS ==========
    headers = ['#', 'Producto', 'Categoría', 'Cantidad', 'Precio Unit.', 'Subtotal']
    header_row = 8
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center
    
    # Datos de productos
    row = header_row + 1
    for idx, detalle in enumerate(detalles, 1):
        ws.cell(row=row, column=1, value=idx).alignment = center
        ws.cell(row=row, column=2, value=detalle.Id_producto.Nombre).alignment = left
        ws.cell(row=row, column=3, value=detalle.Id_producto.Catego_Id.Nombre if detalle.Id_producto.Catego_Id else "N/A").alignment = center
        ws.cell(row=row, column=4, value=detalle.Cantidad).alignment = center
        ws.cell(row=row, column=5, value=float(detalle.Id_producto.Precio_de_venta)).alignment = right
        ws.cell(row=row, column=5).number_format = '"$"#,##0'
        ws.cell(row=row, column=6, value=float(detalle.Subtotal)).alignment = right
        ws.cell(row=row, column=6).number_format = '"$"#,##0'
        
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = border
        
        row += 1
    
    # ========== TOTALES ==========
    row += 1
    subtotal_sin_iva = float(venta.Total) / 1.19
    iva = float(venta.Total) - subtotal_sin_iva
    
    ws.cell(row=row, column=5, value="Subtotal:").font = Font(bold=True)
    ws.cell(row=row, column=5).alignment = right
    ws.cell(row=row, column=6, value=subtotal_sin_iva).number_format = '"$"#,##0'
    ws.cell(row=row, column=6).alignment = right
    
    row += 1
    ws.cell(row=row, column=5, value="IVA (19%):").font = Font(bold=True)
    ws.cell(row=row, column=5).alignment = right
    ws.cell(row=row, column=6, value=iva).number_format = '"$"#,##0'
    ws.cell(row=row, column=6).alignment = right
    
    row += 1
    ws.cell(row=row, column=5, value="TOTAL:").font = Font(bold=True, size=12, color="8B0000")
    ws.cell(row=row, column=5).alignment = right
    ws.cell(row=row, column=6, value=float(venta.Total)).number_format = '"$"#,##0'
    ws.cell(row=row, column=6).font = Font(bold=True, size=12, color="8B0000")
    ws.cell(row=row, column=6).alignment = right
    
    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    
    # Crear respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="Venta_{venta.Id_venta}.xlsx"'
    
    wb.save(response)
    return response


# =============================================
# EXPORTAR A PDF
# =============================================
@login_required(login_url='login')
@permission_required('Ventas.view_venta', login_url='Ventas')
def ExportarVentaPDF(request, id):
    """Exportar detalle de venta a PDF"""
    
    venta = get_object_or_404(Venta.objects.select_related('id_usuario'), Id_venta=id)
    detalles = Detalle_Venta.objects.filter(Id_venta=venta).select_related('Id_producto__Catego_Id')
    
    # Crear buffer para el PDF
    buffer = BytesIO()
    
    # Crear documento PDF
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=50,
        leftMargin=50,
        topMargin=50,
        bottomMargin=50
    )
    
    # Estilos
    styles = getSampleStyleSheet()
    
    # Estilo personalizado para título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=20,
        textColor=colors.HexColor('#8B0000'),
        alignment=TA_CENTER,
        spaceAfter=10
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Heading2'],
        fontSize=14,
        alignment=TA_CENTER,
        spaceAfter=20
    )
    
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontSize=10,
        spaceAfter=5
    )
    
    # Elementos del PDF
    elements = []
    
    # ========== ENCABEZADO ==========
    elements.append(Paragraph("ATHLETIC-Q GIMNASIO", title_style))
    elements.append(Paragraph("Comprobante de Venta", subtitle_style))
    elements.append(Spacer(1, 10))
    
    # ========== INFORMACIÓN DE LA VENTA ==========
    info_data = [
        [Paragraph(f"<b>Venta #:</b> {venta.Id_venta}", normal_style),
         Paragraph(f"<b>Fecha:</b> {venta.Fecha.strftime('%d/%m/%Y %H:%M')}", normal_style)],
        [Paragraph(f"<b>Vendedor:</b> {venta.id_usuario.first_name} {venta.id_usuario.last_name}" if venta.id_usuario else "<b>Vendedor:</b> N/A", normal_style),
         Paragraph(f"<b>Método de Pago:</b> {detalles.first().Tipo_Pago if detalles.first() else 'N/A'}", normal_style)],
    ]
    
    if venta.Cedula_Vents or venta.Numero_Transaccion:
        info_data.append([
            Paragraph(f"<b>Cédula Cliente:</b> {venta.Cedula_Vents or 'N/A'}", normal_style),
            Paragraph(f"<b># Transacción:</b> {venta.Numero_Transaccion or 'N/A'}", normal_style)
        ])
    
    info_table = Table(info_data, colWidths=[250, 250])
    info_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 20))
    
    # ========== TABLA DE PRODUCTOS ==========
    table_data = [['#', 'Producto', 'Categoría', 'Cant.', 'Precio Unit.', 'Subtotal']]
    
    for idx, detalle in enumerate(detalles, 1):
        table_data.append([
            str(idx),
            detalle.Id_producto.Nombre[:30],  # Limitar longitud
            detalle.Id_producto.Catego_Id.Nombre if detalle.Id_producto.Catego_Id else "N/A",
            str(detalle.Cantidad),
            f"${detalle.Id_producto.Precio_de_venta:,.0f}",
            f"${detalle.Subtotal:,.0f}"
        ])
    
    # Crear tabla
    product_table = Table(table_data, colWidths=[30, 150, 90, 40, 80, 80])
    
    # Estilo de la tabla
    product_table.setStyle(TableStyle([
        # Encabezado
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#8B0000')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        
        # Cuerpo
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # Columna #
        ('ALIGN', (3, 1), (3, -1), 'CENTER'),  # Columna Cantidad
        ('ALIGN', (4, 1), (-1, -1), 'RIGHT'),  # Columnas de precios
        
        # Bordes
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        
        # Padding
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        
        # Alternar colores de fila
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
    ]))
    
    elements.append(product_table)
    elements.append(Spacer(1, 20))
    
    # ========== TOTALES ==========
    subtotal_sin_iva = float(venta.Total) / 1.19
    iva = float(venta.Total) - subtotal_sin_iva
    
    totals_data = [
        ['', '', '', '', 'Subtotal:', f"${subtotal_sin_iva:,.0f}"],
        ['', '', '', '', 'IVA (19%):', f"${iva:,.0f}"],
        ['', '', '', '', 'TOTAL:', f"${venta.Total:,.0f}"],
    ]
    
    totals_table = Table(totals_data, colWidths=[30, 150, 90, 40, 80, 80])
    totals_table.setStyle(TableStyle([
        ('FONTNAME', (4, 0), (4, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (4, 0), (-1, -1), 10),
        ('ALIGN', (4, 0), (-1, -1), 'RIGHT'),
        ('TEXTCOLOR', (4, 2), (-1, 2), colors.HexColor('#8B0000')),
        ('FONTNAME', (4, 2), (-1, 2), 'Helvetica-Bold'),
        ('FONTSIZE', (4, 2), (-1, 2), 12),
        ('LINEABOVE', (4, 2), (-1, 2), 1, colors.HexColor('#8B0000')),
    ]))
    
    elements.append(totals_table)
    elements.append(Spacer(1, 40))
    
    # ========== PIE DE PÁGINA ==========
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.grey,
        alignment=TA_CENTER
    )
    
    elements.append(Paragraph("─" * 80, footer_style))
    elements.append(Spacer(1, 10))
    elements.append(Paragraph("Gracias por su compra", ParagraphStyle('Thanks', parent=styles['Normal'], fontSize=12, alignment=TA_CENTER)))
    elements.append(Paragraph("Athletic-Q Gimnasio | www.athleticq.online", footer_style))
    elements.append(Paragraph(f"Documento generado el {timezone.now().strftime('%d/%m/%Y %H:%M')}", footer_style))
    
    # Construir PDF
    doc.build(elements)
    
    # Obtener contenido del buffer
    pdf = buffer.getvalue()
    buffer.close()
    
    # Crear respuesta HTTP
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Venta_{venta.Id_venta}.pdf"'
    response.write(pdf)
    
    return response
